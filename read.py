import pandas as pd
import numpy as np
import openpyxl as xl

data = xl.load_workbook('data.xlsx')
dataframe1 = data.active
for row in range(0, dataframe1.max_row):
    for col in dataframe1.iter_cols(1, dataframe1.max_column):
        print(col[row].value)